from django.shortcuts import render, redirect, get_object_or_404
from dashboard.models import Daily_Profit
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from openpyxl.styles import PatternFill, Font, Alignment
from django.views.generic import ListView, DeleteView
from django.views.decorators.csrf import csrf_exempt
from django.template.loader import get_template
from django.utils.dateparse import parse_date
from django.forms import modelformset_factory
from openpyxl.utils import get_column_letter
from django.views.generic import CreateView
from .models import Order, OrderRequest
from django.http import HttpResponse
from django.http import JsonResponse
from django.urls import reverse_lazy
from django.contrib import messages
from django.utils import timezone
from django.db.models import Sum
from django.db.models import Q
from xhtml2pdf import pisa
from .forms import *
from .utils import *
import openpyxl
import csv


# ------------Order Section Start------------
@csrf_exempt
def update_order(request, pk):
    if request.method == 'POST':
        try:
            order = get_object_or_404(Order, pk=pk)
            data = request.POST
            work_assign_id = data.get('work_assign')
            
            if order.request_order:
                request_order = order.request_order
                request_order.company = data.get('company', request_order.company)
                request_order.name = data.get('name', request_order.name)
                request_order.phone_number = data.get('phone_number', request_order.phone_number)
                request_order.save()  # Save the related model
            elif order.order_customer:
                order_customer = order.order_customer
                order_customer.company = data.get('company', order_customer.company)
                order_customer.name = data.get('name', order_customer.name)
                order_customer.phone_number = data.get('phone_number', order_customer.phone_number)
                order_customer.save()
            
            # Update fields of the Order model
            order.status = data.get('status', order.status)
            order.remark = data.get('remark', order.remark)
            order.delivery_address = data.get('delivery_address', order.delivery_address)
            
            if work_assign_id:
                work_assign_user = get_object_or_404(Custom_User, pk=work_assign_id)
                order.work_assign = work_assign_user
            
            if data.get('delivery_date'):
                order.delivery_date = data.get('delivery_date', order.delivery_date)
            elif not data.get('delivery_date'):
                order.delivery_date = None
            
            order.save()
            
            return JsonResponse({'success': True, 'message': 'Order updated successfully!'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})


class OrderListView(LoginRequiredMixin, ListView):
    model = Order
    template_name = 'order/order.html'
    paginate_by = 20
    context_object_name = 'orders'
    ordering = ['-order_date']
    
    def get_queryset(self):
        queryset = super().get_queryset()

        status = self.request.GET.get('status')
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        search_query = self.request.GET.get('search')
        product_id = self.request.GET.get('product')
        work_assign = self.request.GET.get('work_assign')
        today_orders = self.request.GET.get('today_orders')
        urgent = self.request.GET.get('urgent')
        
        # Filter by urgent
        if urgent == 'true':
            queryset = queryset.filter(urgent=True)

        # Filter by today order
        if today_orders == 'true':
            today = timezone.localtime().date()
            
            queryset = queryset.filter(order_date=today)
            return queryset

        # Filter by status
        if status and status != 'All':
            queryset = queryset.filter(status=status)

        # Filter by date range
        if start_date:
            start_date = parse_date(start_date)
            queryset = queryset.filter(order_date__gte=start_date)
        if end_date:
            end_date = parse_date(end_date)
            queryset = queryset.filter(order_date__lte=end_date)
        

        # Filter by Search
        if search_query:
            queryset = queryset.filter(
                Q(tracking_ID__icontains=search_query) |
                Q(request_order__name__icontains=search_query) |
                Q(request_order__company__icontains=search_query) |
                Q(request_order__phone_number__icontains=search_query) |
                Q(request_order__product__name__icontains=search_query) |
                Q(order_customer__name__icontains=search_query) |
                Q(order_customer__company__icontains=search_query) |
                Q(order_customer__phone_number__icontains=search_query) |
                Q(order_customer__product__name__icontains=search_query)
            )
            
        
        # Filter by Product
        if product_id and product_id.isdigit():
            queryset = queryset.filter(
                Q(request_order__product__id__icontains=product_id) |
                Q(order_customer__product__id__icontains=product_id)
            ).distinct()
            
        if work_assign and work_assign.isdigit():
            queryset = queryset.filter(work_assign__id__icontains=work_assign).distinct()

        return queryset.distinct()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['products'] = Product.objects.all()
        context['work_assign_choices'] = Custom_User.objects.all()
        context['users'] = Custom_User.objects.all()
        
        queryset = self.get_queryset()
        context['total_deal_value'] = queryset.aggregate(Sum('deal_value'))['deal_value__sum'] or 0
        context['total_due_amount'] = queryset.aggregate(Sum('due_amount'))['due_amount__sum'] or 0
        context['total_quantity'] = sum(sum(item.quantity for item in order.order_item.all()) for order in queryset)
        
        # context = {
        #     "order_data": json.dumps([
        #         [
        #             order.order_date,
        #             order.tracking_ID,
        #             order.request_order.company if order.request_order else order.order_customer.company,
        #             order.request_order.name if order.request_order else order.order_customer.name,
        #             ", ".join([p.name for p in (order.request_order.product.all() if order.request_order else order.order_customer.product.all())]),
        #             order.work_assign,
        #             order.status,
        #             order.delivery_date,
        #             order.deal_value,
        #             order.due_amount,
        #             order.pathao_parcel_id if order.pathao_parcel_id else "No",
        #             "View/Delete",  # Placeholder for actions
        #         ]
        #         for order in queryset
        #     ]),
        # }
        
        return context
    
    
    def render_to_response(self, context, **response_kwargs):
        export_format = self.request.GET.get('export')
        queryset = self.get_queryset()
        headers = [
            "Tracking ID", "Company", "Customer Name", "Phone Number", "Second Phone Number", "Email",
            "Delivery Address", "Special Instructions", "Source", "Products", "Assigned Work",
            "Status", "Remark", "Logo", "Picture1", "Picture2", "Picture3", "Picture4", "Picture5",
            "Product", "Quantity", "Unit Price", "Item Total", "Deal Value", "Delivery Charge", "Advance Amount", "Due Amount",
            "Delivery Charge Cost", "Extra Cost", "Total Product Amount","Payment Status", "Payment Method", "Payment Number", "Transaction ID", "Delivery Date", "Last Update", "Order Date"
        ]
        
        if export_format == 'xlsx':
            return self.export_to_xlsx(queryset, headers)
        elif export_format == 'csv':
            return self.export_to_csv(queryset, headers)

        return super().render_to_response(context, **response_kwargs)
    
    
    def export_to_xlsx(self, orders, headers):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Orders'
        sheet.append(headers)
        
        header_fill = PatternFill(start_color="346754", end_color="346754", fill_type="solid")  # Yellow background
        header_font = Font(color="ffffff", bold=True)  # Red font, bold
        
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        column_widths = [
            12, 20, 20, 15, 20, 20,   # Tracking ID, Company, Customer Name, Phone Number, Second Phone Number, Email
            25, 20, 12, 25, 15,   # Delivery Address, Special Instructions, Source, Products, Assigned Work
            10, 15, 15, 10, 10,   # Status, Remark, Logo, Picture1, Picture2
            10, 10, 10, 15, 10, 10,   # Picture3, Picture4, Picture5, Product, Quantity, Unit Price
            10, 10, 15, 15, 12,   # Item Total, Deal Value, Delivery Charge, Advance Amount, Due Amount
            18, 10, 20, 15, 15,   # Delivery Charge Cost, Extra Cost, Total Amount, Payment Number, Transaction ID
            15, 15, 15, 15, 15         # Payment Method, Payment Status, Delivery Date, Last Update, Order Date
        ]

        for i, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(i)
            sheet.column_dimensions[column_letter].width = width

        current_row = 2

        # Populate data rows
        for order in orders:
            if order.request_order:
                customer = order.request_order
            else:
                customer = order.order_customer

            if not customer:
                continue

            products = ', '.join([product.name for product in customer.product.all()])

            # Get order items and calculate row range to merge cells over
            order_items = list(order.order_item.all())
            start_row = current_row
            end_row = start_row + len(order_items) - 1

            # Fill data for each order item, only showing order-level fields on the first row
            for idx, item in enumerate(order_items):
                row = []

                # Only display order-level fields on the first row for each order
                if idx == 0:
                    row += [
                        order.tracking_ID,
                        customer.company,
                        customer.name,
                        customer.phone_number,
                        customer.second_phone_number,
                        customer.email,
                        order.delivery_address,
                        order.special_instructions,
                        customer.source,
                        products,
                        str(order.work_assign),
                        order.status,
                        order.remark,
                        customer.logo,
                        customer.picture1,
                        getattr(customer, 'picture2', None),
                        getattr(customer, 'picture3', None),
                        getattr(customer, 'picture4', None),
                        getattr(customer, 'picture5', None)
                    ]
                else:
                    row += [''] * 19

                # Item-specific fields
                row += [
                    item.product.name,
                    item.quantity,
                    item.unit_price,
                    item.total
                ]

                # Order totals and other fields (only on the first row of the order)
                if idx == 0:
                    row += [
                        order.deal_value,
                        order.delivery_charge,
                        order.advance_amount,
                        order.due_amount,
                        order.delivery_charge_cost,
                        order.extra_cost,
                        order.total_amount,
                        order.payment_number,
                        order.transaction_id,
                        order.payment_method,
                        order.payment_status,
                        order.delivery_date.strftime('%Y-%m-%d') if order.delivery_date else '',
                        order.last_update.strftime('%Y-%m-%d'),
                        order.order_date.strftime('%Y-%m-%d')
                    ]
                else:
                    row += [''] * 15

                sheet.append(row)
                for col_num, value in enumerate(row, start=1):
                    cell = sheet.cell(row=current_row, column=col_num, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")  # Center align all cells
                current_row += 1

            # Merge cells for each order-level field across the rows for this order
            if len(order_items) > 1:
                merge_columns = list(range(1, 20)) + list(range(24, 38))  # Columns to merge (tracking ID, company, etc.)
                for col_idx in merge_columns:
                    col_letter = get_column_letter(col_idx)
                    sheet.merge_cells(f"{col_letter}{start_row}:{col_letter}{end_row}")
                    # Center align merged cells
                    merged_cell = sheet.cell(row=start_row, column=col_idx)
                    merged_cell.alignment = Alignment(horizontal="center", vertical="center")

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Orders - {timezone.now().date()}.xlsx'
        workbook.save(response)
        return response

    def export_to_csv(self, orders, headers):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename=orders - {timezone.now().date()}.csv'
        writer = csv.writer(response)
        writer.writerow(headers)

        # Populate data rows
        for order in orders:
            if order.request_order:
                customer = order.request_order
            else:
                customer = order.order_customer

            if not customer:
                continue

            products = ', '.join([product.name for product in customer.product.all()])

            # Iterate over order items and only display deal value and other order-level fields on the first row
            order_items = list(order.order_item.all())
            for idx, item in enumerate(order_items):
                row = []

                # Display order-level fields only for the first row of this order
                if idx == 0:
                    row += [
                        order.tracking_ID,
                        customer.company,
                        customer.name,
                        customer.phone_number,
                        customer.second_phone_number,
                        customer.email,
                        order.delivery_address,
                        order.special_instructions,
                        customer.source,
                        products,
                        str(order.work_assign),
                        order.status,
                        order.remark,
                        customer.logo,
                        customer.picture1,
                        getattr(customer, 'picture2', None),
                        getattr(customer, 'picture3', None),
                        getattr(customer, 'picture4', None),
                        getattr(customer, 'picture5', None)
                    ]
                else:
                    # Add empty cells for these columns in subsequent rows for the same order
                    row += [''] * 19

                # Add item-specific fields for each row
                row += [
                    item.product.name,
                    item.quantity,
                    item.unit_price,
                    item.total
                ]

                # Display these only for the first item of the order
                if idx == 0:
                    row += [
                        order.deal_value,
                        order.delivery_charge,
                        order.advance_amount,
                        order.due_amount,
                        order.delivery_charge_cost,
                        order.extra_cost,
                        order.total_amount,
                        order.payment_status,
                        order.payment_method,
                        order.payment_number,
                        order.transaction_id,
                        order.delivery_date.strftime('%Y-%m-%d') if order.delivery_date else '',
                        order.last_update.strftime('%Y-%m-%d'),
                        order.order_date.strftime('%Y-%m-%d')
                    ]
                else:
                    # Append empty fields for subsequent items
                    row += [''] * 15

                writer.writerow(row)

        return response
    

class OrderDeleteView(LoginRequiredMixin, DeleteView):
    model = Order
    template_name = 'order/order_confirm_delete.html'
    success_url = reverse_lazy('order_list')


@login_required
def add_new_order(request):
    if request.method == 'POST':
        order_customer_form = OrderCustomerForm(request.POST)
        order_form = OrderForm(request.POST, request.FILES)

        if order_customer_form.is_valid() and order_form.is_valid():
            order_customer = order_customer_form.save()
            order = order_form.save(commit=False)
            order.order_customer = order_customer
            order.save()
            return redirect('order_success', id=order.id)
        else:
            messages.warning(
                request, f"{order_customer_form.errors} and {order_form.errors}")
            return redirect(request.META['HTTP_REFERER'])
    else:
        order_customer_form = OrderCustomerForm()
        order_form = OrderForm()

    return render(request, 'order/add_new_order.html', {
        'order_customer_form': order_customer_form,
        'order_form': order_form
    })


@login_required
def order_success(request, id):
    order = get_object_or_404(Order, id=id)
    return render(request, 'order/order_success.html', {'order': order})


@login_required
def DeleteOrderItem(request, id):
    order_item = get_object_or_404(OrderItem, id=id)
    order = order_item.order_items.all().first()
    order.order_item.remove(order_item)
    if order.request_order:
        order.request_order.product.remove(order_item.product)
    elif order.order_customer:
        order.order_customer.product.remove(order_item.product)
    order_item.delete()
    return redirect(request.META['HTTP_REFERER'])


@login_required
def order_view(request, id):
    order = get_object_or_404(Order, id=id)
    OrderItemFormSet = modelformset_factory(OrderItem, form=OrderItemUpdateForm, extra=0)

    if request.method == 'POST':
        form = OrderStatusUpdateForm(request.POST, request.FILES, instance=order)
        if form.is_valid():
            form.save()
            messages.success(request, 'Order updated successfully!')
            return redirect(request.META['HTTP_REFERER'])
        else:
            messages.error(request, f'Something is wrong: {form.errors} {item_formset.errors}')
            return redirect(request.META['HTTP_REFERER'])
    else:
        form = OrderStatusUpdateForm(instance=order)
        form2 = OrderPaymentUpdateForm(instance=order)
        item_formset = OrderItemFormSet(queryset=order.order_item.all())
        context = {
            'order': order,
            'form': form,
            'form2': form2,
            'item_formset': item_formset,
        }
        if order.order_customer and order.request_order is None:
            form3 = OrderCustomerForm(instance=order.order_customer)
            form3.fields['product'].required = False
            context['form3'] = form3

    return render(request, 'order/order_view.html', context)


@login_required
def create_pathao_parcel(request, id):
    order = get_object_or_404(Order, id=id)
    if order.request_order:
        recipient = order.request_order
    elif order.order_customer:
        recipient = order.order_customer
    
    access_token = get_access_token()
    pathao_order_data = {
            'store_id':"90968",
            'merchant_order_id':order.tracking_ID,
            'recipient_name':recipient.name,
            'recipient_phone':recipient.phone_number,
            'recipient_address':order.delivery_address,
            # 'recipient_city':'1',
            # 'recipient_zone':'4',
            # 'recipient_area':'105',
            'delivery_type':48,
            'item_type':2,
            'special_instruction':order.special_instructions,
            'item_quantity':sum(item.quantity for item in order.order_item.all()),
            'item_weight':1,
            'amount_to_collect':float(order.due_amount),
            'item_description':order.remark
        }
    pathao_response = create_pathao_order(pathao_order_data, access_token)
    if pathao_response['type'] == 'success':
        consignment_id = pathao_response.get("data", {}).get("consignment_id")
        if consignment_id:
            order.pathao_parcel_id = consignment_id
            order.status = 'Delivered'
            order.save()
        return redirect(request.META['HTTP_REFERER'])
    
    elif pathao_response['type'] == 'error':
        error_message = ""
        for field, message in pathao_response['errors'].items():
            for message in message:
                error_message += f"{message}"
        
        # for field, message in pathao_response['errors'].items():
        #     error_message += f"{field}:\n"
        #     for message in message:
        #         error_message += f"{message}\n"
        
        messages.error(request, error_message)
        return redirect(request.META['HTTP_REFERER'])
    else:
        return redirect(request.META['HTTP_REFERER'])


def render_to_pdf(template_src, context_dict={}):
    template = get_template(template_src)
    html = template.render(context_dict)
    response = HttpResponse(content_type='application/pdf')
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response


@login_required
def print_invoice(request, id):
    order = get_object_or_404(Order, id=id)
    context = {'order': order}
    return render_to_pdf('order/invoice.html', context)


@login_required
def orderPaymentUpdate(request, pk):
    order = get_object_or_404(Order, pk=pk)
    OrderItemFormSet = modelformset_factory(OrderItem, form=OrderItemUpdateForm, extra=0)
    
    if request.method == 'POST':
        form2 = OrderPaymentUpdateForm(request.POST, instance=order)
        item_formset = OrderItemFormSet(request.POST, queryset=order.order_item.all())

        if form2.is_valid() and item_formset.is_valid():
            item_formset.save()
            order = form2.save()
            # daily_profit_update(order)
            
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':  # AJAX request check
                return JsonResponse({'success': True, 'message': 'Payment details updated successfully!'})

            messages.success(request, 'Payment details updated successfully!')
            return redirect('order_view', id=order.pk)
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form2.errors, 'item_errors': item_formset.errors})
            
            messages.error(request, f'Something went wrong: {form2.errors} {item_formset.errors}')
            return redirect('order_view', id=order.pk)


@login_required
def CustomerOrderInfoUpdate(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if request.method == 'POST':
        form3 = OrderCustomerForm(request.POST, instance=order.order_customer)
        if form3.is_valid():
            form3.save()
            messages.success(request, 'Order Customer Info updated successfully!')
        else:
            messages.error(request, f'Something went wrong: {form3.errors}')
    return redirect('order_view', id=order.pk)

# ------------Order Section End------------




# ------------Order Request Section Start------------
class OrderRequestListView(LoginRequiredMixin, ListView):
    model = OrderRequest
    template_name = 'request/order_request_list.html'
    context_object_name = 'order_requests'
    paginate_by = 20  # Pagination works for normal listing
    ordering = ['-id']

    def get_queryset(self):
        queryset = super().get_queryset()

        # Get filter parameters from the request
        status = self.request.GET.get('status')
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        search_query = self.request.GET.get('search')
        product_id = self.request.GET.get('product')
        work_assign = self.request.GET.get('work_assign')
        today_order_requests = self.request.GET.get('today_order_requests')
        urgent = self.request.GET.get('urgent')
        
        if urgent == 'true':
            queryset = queryset.filter(urgent=True)

        # Apply filters
        if today_order_requests == 'true':
            today = timezone.localtime().date()
            queryset = queryset.filter(created_at=today)

        if status and status != 'All':
            queryset = queryset.filter(status=status)

        if start_date:
            start_date = parse_date(start_date)
            queryset = queryset.filter(created_at__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            queryset = queryset.filter(created_at__lte=end_date)

        if search_query:
            queryset = queryset.filter(
                Q(tracking_ID__icontains=search_query) |
                Q(name__icontains=search_query) |
                Q(company__icontains=search_query) |
                Q(phone_number__icontains=search_query) |
                Q(source__icontains=search_query) |
                Q(status__icontains=search_query) |
                Q(remark__icontains=search_query) |
                Q(product__name__icontains=search_query)
            )

        if product_id and product_id.isdigit():
            queryset = queryset.filter(product__id__icontains=product_id).distinct()
            
        if work_assign and work_assign.isdigit():
            queryset = queryset.filter(work_assign__id__icontains=work_assign).distinct()
        
        
        return queryset.distinct()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['products'] = Product.objects.all()
        context['work_assign_choices'] = Custom_User.objects.all()
        return context

    def render_to_response(self, context, **response_kwargs):
        export_format = self.request.GET.get('export')
        queryset = self.get_queryset()  # Get the filtered queryset without pagination
        headers = [
            'Created By', 'ID', 'Company', 'Name', 'Phone Number', 'Second Phone Number', 'email', 'Source', 'Product(s)', 'Status', 'remark', 'Work Assign', 'Order Created', 'Last Update', 'Created At', 'logo', 'picture1', 'picture2', 'picture3', 'picture4', 'picture5'
        ]
        
        if export_format == 'xlsx':
            return self.export_to_xlsx(queryset, headers)
        elif export_format == 'csv':
            return self.export_to_csv(queryset, headers)

        return super().render_to_response(context, **response_kwargs)

    def export_to_xlsx(self, order_requests, headers):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f'Order Requests - {timezone.now().date()}'
        sheet.append(headers)
        
        header_fill = PatternFill(start_color="346754", end_color="346754", fill_type="solid")  # Yellow background
        header_font = Font(color="ffffff", bold=True)  # Red font, bold
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        column_widths = [
            15, 12, 20, 20, 15, 20, 20,   # Created By, ID, Company, Name, Phone Number, second_phone_number, Email
            12, 25, 10, 20, 15,   # Source, Products, Status, remark, Assigned Work
            10, 15, 15, # Order Created, Last Update, Created At, 
            15, 10, 10, 10, 10, 10  # Logo, Picture1, Picture2, Picture3, Picture4, Picture5
        ]
        for i, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(i)
            sheet.column_dimensions[column_letter].width = width
            

        for order_request in order_requests:
            products = ', '.join([product.name for product in order_request.product.all()])
            sheet.append([
                str(order_request.created_by),
                order_request.tracking_ID,
                order_request.company,
                order_request.name,
                order_request.phone_number,
                order_request.second_phone_number,
                order_request.email,
                order_request.source,
                products,
                order_request.status,
                order_request.remark,
                str(order_request.work_assign),
                order_request.order_created,
                str(order_request.last_update),
                str(order_request.created_at),
                order_request.logo,
                order_request.picture1,
                order_request.picture2,
                order_request.picture3,
                order_request.picture4,
                order_request.picture5
            ])
            # Center align merged cells
            # merged_cell = sheet.cell(row=1, column=1)
            # merged_cell.alignment = Alignment(horizontal="center", vertical="center")

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=order_requests - {timezone.now().date()}.xlsx'
        workbook.save(response)
        return response

    def export_to_csv(self, order_requests, headers):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename=order_requests - {timezone.now().date()}.csv'
        writer = csv.writer(response)
        writer.writerow(headers)

        for order_request in order_requests:
            products = ', '.join([product.name for product in order_request.product.all()])
            writer.writerow([
                order_request.created_by,
                order_request.tracking_ID,
                order_request.company,
                order_request.name,
                order_request.phone_number,
                order_request.second_phone_number,
                order_request.email,
                order_request.source,
                products,
                order_request.status,
                order_request.remark,
                order_request.work_assign,
                order_request.order_created,
                order_request.last_update,
                order_request.created_at,
                order_request.logo,
                order_request.picture1,
                order_request.picture2,
                order_request.picture3,
                order_request.picture4,
                order_request.picture5
            ])

        return response


class OrderRequestCreateView(LoginRequiredMixin, CreateView):
    model = OrderRequest
    form_class = OrderRequestForm
    template_name = 'request/order_request_form.html'
    success_url = reverse_lazy('order_request_list')


class OrderRequestDeleteView(LoginRequiredMixin, DeleteView):
    model = OrderRequest
    template_name = 'request/order_request_confirm_delete.html'
    success_url = reverse_lazy('order_request_list')


@login_required
def order_request_view(request, pk):
    order_request = get_object_or_404(OrderRequest, id=pk)

    if request.method == 'POST':
        form = OrderRequestStatusUpdateForm(request.POST, instance=order_request)
        if form.is_valid():
            form.save()
            if hasattr(order_request, 'order'):
                order_request.order.save()
        else:
            messages.error(request, f'Something went wrong: {form.errors}')
        return redirect('order_request_view', pk=order_request.pk)
    else:
        form = OrderRequestStatusUpdateForm(instance=order_request)
        form2 = OrderForm()
        form3 = OrderRequestPictureUpdateForm(instance=order_request)

    return render(request, 'request/order_request_view.html', {'order_request': order_request, 'form': form, 'form2': form2, 'form3': form3})


@login_required
def request_to_order(request, pk):
    if request.method == 'POST':
        order_request = get_object_or_404(OrderRequest, id=pk)
        if Order.objects.filter(tracking_ID=order_request.tracking_ID).exists():
            messages.success(request, "Order already created!")
            return redirect('order_request_view', pk=order_request.pk)
        else:
            form2 = OrderForm(request.POST, request.FILES)
            if form2.is_valid():
                order = form2.save()
                order.request_order = order_request
                # order.tracking_ID = order_request.tracking_ID
                if order.remark is None:
                    order.remark = order_request.remark
                order.save()

                order_request.order_created = True
                order_request.save()

                messages.success(request, 'Order created successfully.')
                return redirect('order_view', id=order.id)
            else:
                messages.success(request, form2.errors)
                return redirect('order_request_view', pk=order_request.pk)


@login_required
def PictureUpdate(request, pk):
    order_request = get_object_or_404(OrderRequest, pk=pk)
    if request.method == 'POST':
        form3 = OrderRequestPictureUpdateForm(request.POST, instance=order_request)
        if form3.is_valid():
            form3.save()
            messages.success(request, 'Picture details updated successfully!')
        else:
            messages.error(request, f'Something went wrong: {form3.errors}')
    return redirect('order_request_view', pk=order_request.pk)

# ------------Order Request Section End------------
